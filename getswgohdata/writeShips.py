import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook

def writeLevels(toonDictionary, toonNames, memberNames, sheet):
    characterCounter = 0
    memberCounter = 0
    for k, v in toonDictionary.items():
        for i, j in v.items():
            print(str(k) + '  ' +  str(i) + ' ' + str(j))
            for m in range(len(toonNames)):
                if(i == toonNames[m]):
                    characterCounter = m
            sheet.cell(row=(characterCounter+2), column=(memberCounter+2)).value = j
        memberCounter += 1
        characterCounter = 0
    for i in range(len(toonNames)):
        for j in range(len(memberNames)):
            if(sheet.cell(row=i+2, column=j+2).value == None):
                sheet.cell(row=i+2, column=j+2).value = 0

def getWidth(stringList):
    maxCharacterCount = 0
    for i in range(len(stringList)):
        currentString = stringList[i]
        #print(stringList[i] + " length is " + str(len(currentString)))
        if(len(currentString) >= maxCharacterCount):
            maxCharacterCount = len(currentString)
    return maxCharacterCount

def writeMemberNames(stringList, sheet):
    dimension = getWidth(stringList)
    for i in range(len(stringList)):
        sheet.cell(row=1, column=(i+2)).value = stringList[i]
        sheet.cell(row=1, column=(i+2)).alignment = Alignment(horizontal='right', text_rotation=90)
        sheet.row_dimensions[1].height = (7*dimension)
        sheet.column_dimensions[get_column_letter(i+2)].width = 3

def writeCharacterNames(stringList, sheet):
    columnWidth = getWidth(stringList)
    stringList.sort()
    for i in range(len(stringList)):
        sheet.cell(row=i+2, column=1).value = stringList[i]
        sheet.column_dimensions[get_column_letter(1)].width = columnWidth
        

def getCharacterNames(toonsDictionary):
    toonsName = []
    for k, v in toonsDictionary.items():
        for i, j in v.items():
            if(i in toonsName):
                continue
            else:
                toonsName.append(i)
    return toonsName
    
def write(memberNames, memberShips):
    wb = load_workbook('example.xlsx')
    shipNames = []
    sheet = wb.create_sheet('Ships')
    shipNames = getCharacterNames(memberShips)
    writeMemberNames(memberNames, sheet)
    writeCharacterNames(shipNames, sheet)
    writeLevels(memberShips, shipNames, memberNames, sheet)
    wb.save('example.xlsx')
