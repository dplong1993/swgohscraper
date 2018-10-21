# This program is a class that represents
# a swgoh guild.

import requests, bs4
import member
import completeNameRosterForToons
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook


class Guild:
    # TODO MAKE THIS LOOK LIKE MEMBER init FUNCTION!!!!!
    def __init__(self, url):
        # Set the guild's url.
        self.__guild_url = url

        # Use url to get html of guild from SWGOH.gg
        res = requests.get(self.__guild_url)
        soup = bs4.BeautifulSoup(res.text, 'html.parser')

        # Set the guild's member list.
        self.__members = []
        member_tag = soup.body.table.tbody.tr  # This is the tag for the first member of the guild

        # Sort through the html to get all the guild's members
        while member_tag is not None:
            # Get the member's name and link from the html.
            name = str(member_tag.a.strong.string)
            link = member_tag.a.get('href')

            # Create a member object for that member.
            print("Adding: " + name)
            new_member = member.Member(name, link)

            # Add the member to the Guild's member list.
            self.__members.append(new_member)

            # Go to the next member in the html.
            member_tag = member_tag.next_sibling.next_sibling

        # Set the guild's member count.
        self.__member_count = len(self.__members)

    # Method to return a list of all member names. (debug checked)
    def get_list_of_member_names(self):
        member_list = self.get_list_of_member_objects()
        members_name_list = []
        for i in member_list:
            members_name_list.append(i.get_name())
        return members_name_list

    # Method to print list of all member's names and urls and gps (debugged)
    def print_members_names_urls_and_gps(self):
        member_list = self.get_list_of_member_objects()
        for i in member_list:
            print(str(i))

    # Method to print list of all member's names only (debugged)
    def print_members_names(self):
        member_list = self.get_list_of_member_objects()
        for i in member_list:
            print(i.get_name())

    # Method to get a list of all the guild's member objects. (debug checked)
    def get_list_of_member_objects(self):
        return self.__members

    # Method to get the guild's url. (debugged)
    def get_url(self):
        return self.__guild_url

    # Method to get the number of members in the guild. (debugged)
    def get_member_count(self):
        return self.__member_count

    # Method to get a list of each member's url (debugged)
    def get_members_url_list(self):
        member_list = self.get_list_of_member_objects()
        url_list = []
        for i in member_list:
            url_list.append(i.get_url())
        return url_list

    # Method to print a list of each member's url (testing)
    def print_list_of_members_url(self):
        member_list = self.get_list_of_member_objects()
        for i in member_list:
            print(i.get_url())
        return

    # TODO Method to get a specific member's ls characters
    '''def get_member_light_side_characters(self, name):
        member_list = self.get_members()
        for i in member_list:
            if i.get_name() is name:
                get_light_side_toons_list(i)
        print("That name was not found in the member list for this guild.")
    '''

    # METHOD TO GET A LIST OF THE NAMES OF A SPECIFIC MEMBER'S CHARACTERS AND SHIPS (DONE)
    def get_list_of_a_members_ships_and_characters(self, name):
        name = name.lower()
        member_list = self.get_list_of_member_objects()
        new_list = []
        for i in member_list:
            if name == i.get_name().lower():
                new_list = i.get_list_of_ships_and_characters()
                return new_list
        print("Name is not in the member list. A blank list is being returned.")
        return []

    # METHOD TO PRINT A LIST OF THE NAMES OF A SPECIFIC MEMBER'S CHARACTERS AND SHIPS (DONE)
    def print_list_of_a_members_ships_and_characters(self, name):
        name = name.lower()
        member_list = self.get_list_of_member_objects()
        new_list = []
        for i in member_list:
            if name == i.get_name().lower():
                i.print_list_of_ships_and_characters()
                return
        print("Name is not in the member list. Nothing will be printed.")
        return

    # TODO Return the star level of one member's character or ship based on the name provided

    # TODO Print the star level of one member's character or ship based on the name provided

    # TODO Return the gp level of one member's character or ship based on the name provided

    # TODO Print the gp level of one member's character or ship based on the name provided

    # Return a dictionary of all member names and gp values for a specific character or ship name(DONE)
    def find_units_power_level(self, toon_name):
        member_list = self.get_list_of_member_objects()
        toon_power_dict = {}
        for i in member_list:
            power = i.get_toon_power(toon_name)
            toon_power_dict.setdefault(str(i.get_name()), str(power))
        return toon_power_dict

    # Print a table of all member names and gp values for a specific character or ship name
    def print_units_power_level(self, toon_name):
        member_list = self.get_list_of_member_objects()
        for i in member_list:
            power = i.get_toon_power(toon_name)
            # TODO WORK ON FORMATTING THE PRINT STATEMENT TO MAKE THE TABLE NEAT
            print(str(i.get_name()) + ": " + str(power))

    # Return a dictionary of all member names and star values for a specific character or ship name
    def find_units_star_level(self, toon_name):
        member_list = self.get_list_of_member_objects()
        toon_star_dict = {}
        for i in member_list:
            star = i.get_star_level(toon_name)
            toon_star_dict.setdefault(str(i.get_name()), str(star))
        return toon_star_dict

    # Print a table of all member names and gp values for a specific character or ship name
    def print_units_star_level(self, toon_name):
        member_list = self.get_list_of_member_objects()
        for i in member_list:
            star = i.get_star_level(toon_name)
            # TODO WORK ON FORMATTING THE PRINT STATEMENT TO MAKE THE TABLE NEAT
            print(str(i.get_name()) + ": " + str(star))

    # Print out Member Name, Total GP, Char Gp, and Ship Gp for each member
    def print_member_info(self):
        member_list = self.get_list_of_member_objects()
        for i in member_list:
            name = i.get_name()
            total_power = i.get_total_gp()
            char_power = i.get_char_gp()
            ship_power = i.get_ships_gp()
            print(str(name) + ":" + str(total_power) + ":" + str(char_power) + ":" + str(ship_power))

    # TODO write member names and their character's star levels to excel file.
    def write_member_rosters_to_spreadsheet(self):
        wb = load_workbook('example.xlsx')
        dark_toons_names = []
        member_list = self.get_list_of_member_objects()
        roster = completeNameRosterForToons.CharacterNameList()
        names_list = roster.get_dark_side_names_list()
        sheet = wb.create_sheet('DSToons_Roster')
        for i in range(len(names_list)):
            for j in range(len(member_list)):
                sheet.cell(row=i + 2, column=1).value = names_list[i]
                name = member_list[j].get_name()
                sheet.cell(row=1, column=(j + 2)).value = name
                star_level = member_list[j].get_star_level(names_list[i])
                sheet.cell(row= i + 2, column = j + 2).value = star_level
        sheet = wb.create_sheet('DSToons_Power_Level')
        for i in range(len(names_list)):
            for j in range(len(member_list)):
                sheet.cell(row=i + 2, column=1).value = names_list[i]
                name = member_list[j].get_name()
                sheet.cell(row=1, column=(j + 2)).value = name
                power_level = member_list[j].get_toon_power(names_list[i])
                sheet.cell(row=i + 2, column=j + 2).value = power_level
        sheet = wb.create_sheet('DSToons_Level')
        for i in range(len(names_list)):
            for j in range(len(member_list)):
                sheet.cell(row=i + 2, column=1).value = names_list[i]
                name = member_list[j].get_name()
                sheet.cell(row=1, column=(j + 2)).value = name
                unit_level = member_list[j].get_unit_levels(names_list[i])
                sheet.cell(row=i + 2, column=j + 2).value = unit_level
        sheet = wb.create_sheet('DSToons_Gear_Level')
        for i in range(len(names_list)):
            for j in range(len(member_list)):
                sheet.cell(row=i + 2, column=1).value = names_list[i]
                name = member_list[j].get_name()
                sheet.cell(row=1, column=(j + 2)).value = name
                gear_level = member_list[j].get_gear_level(names_list[i])
                sheet.cell(row=i + 2, column=j + 2).value = gear_level
        names_list = roster.get_light_side_names_list()
        sheet = wb.create_sheet('LSToons_Roster')
        for i in range(len(names_list)):
            for j in range(len(member_list)):
                sheet.cell(row=i + 2, column=1).value = names_list[i]
                name = member_list[j].get_name()
                sheet.cell(row=1, column=(j + 2)).value = name
                star_level = member_list[j].get_star_level(names_list[i])
                sheet.cell(row=i + 2, column=j + 2).value = star_level
        sheet = wb.create_sheet('LSToons_Power_Level')
        for i in range(len(names_list)):
            for j in range(len(member_list)):
                sheet.cell(row=i + 2, column=1).value = names_list[i]
                name = member_list[j].get_name()
                sheet.cell(row=1, column=(j + 2)).value = name
                power_level = member_list[j].get_toon_power(names_list[i])
                sheet.cell(row=i + 2, column=j + 2).value = power_level
        sheet = wb.create_sheet('LSToons_Level')
        for i in range(len(names_list)):
            for j in range(len(member_list)):
                sheet.cell(row=i + 2, column=1).value = names_list[i]
                name = member_list[j].get_name()
                sheet.cell(row=1, column=(j + 2)).value = name
                unit_level = member_list[j].get_unit_levels(names_list[i])
                sheet.cell(row=i + 2, column=j + 2).value = unit_level
        sheet = wb.create_sheet('LSToons_Gear_Level')
        for i in range(len(names_list)):
            for j in range(len(member_list)):
                sheet.cell(row=i + 2, column=1).value = names_list[i]
                name = member_list[j].get_name()
                sheet.cell(row=1, column=(j + 2)).value = name
                gear_level = member_list[j].get_gear_level(names_list[i])
                sheet.cell(row=i + 2, column=j + 2).value = gear_level

        wb.save('example.xlsx')

    # TODO UNFINISHED METHODS
    # Method to print a member's toon roster.
    # @staticmethod
    # def print_member_roster(guild_member):
        # print(35 * '*')
        # print('{:30}'.format(guild_member.get_name()) + '{:>5}'.format('**'))
        # print(35 * '*')
        # guild_member.get_member_toon_roster()

    # Method to print all members and their toon rosters.
    # def print_members_rosters(self):
        # for i in self.__members:
            # self.print_member_roster(i)

    # def get_count_of_character(self, name, stars, alignment):
        # count = 0
        # for i in self.__members:
            # increment = i.get_count_of_character(name, stars, alignment)
            # count += increment
        # return count

    # Displays the guild url followed by the guild's member count
    # whenever the guild object is called in a string
    # def __str__(self):
        # return self.__guild_url + " Member Count: " + str(self.__member_count)
