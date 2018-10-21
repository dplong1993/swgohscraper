from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook


def write_levels(toon_dictionary, toon_names, member_names, sheet):
    character_counter = 0
    member_counter = 0
    for k, v in toon_dictionary.items():
        for i, j in v.items():
            print(str(k) + '  ' + str(i) + ' ' + str(j))
            for m in range(len(toon_names)):
                if i == toon_names[m]:
                    character_counter = m
            sheet.cell(row=(character_counter + 2), column=(member_counter + 2)).value = j
        member_counter += 1
        character_counter = 0
    for i in range(len(toon_names)):
        for j in range(len(member_names)):
            if sheet.cell(row=i + 2, column=j + 2).value is None:
                sheet.cell(row=i + 2, column=j + 2).value = 0


def get_width(string_list):
    max_character_count = 0
    for i in range(len(string_list)):
        current_string = string_list[i]
        # print(stringList[i] + " length is " + str(len(current_string)))
        if len(current_string) >= max_character_count:
            max_character_count = len(current_string)
    return max_character_count


def write_member_names(string_list, sheet):
    dimension = get_width(string_list)
    for i in range(len(string_list)):
        sheet.cell(row=1, column=(i + 2)).value = string_list[i]
        sheet.cell(row=1, column=(i + 2)).alignment = Alignment(horizontal='right', text_rotation=90)
        sheet.row_dimensions[1].height = (7 * dimension)
        sheet.column_dimensions[get_column_letter(i + 2)].width = 3


def write_character_names(string_list, sheet):
    column_width = get_width(string_list)
    string_list.sort()
    for i in range(len(string_list)):
        sheet.cell(row=i + 2, column=1).value = string_list[i]
        sheet.column_dimensions[get_column_letter(1)].width = column_width


def get_character_names(toons_dictionary):
    toons_name = []
    for k, v in toons_dictionary.items():
        for i, j in v.items():
            if i in toons_name:
                continue
            else:
                toons_name.append(i)
    return toons_name


def write(member_names, member_ds_toons):
    wb = load_workbook('example.xlsx')
    dark_toons_names = []
    sheet = wb.create_sheet('DSToons')
    dark_toons_names = get_character_names(member_ds_toons)
    write_member_names(member_names, sheet)
    write_character_names(dark_toons_names, sheet)
    write_levels(member_ds_toons, dark_toons_names, member_names, sheet)
    wb.save('example.xlsx')
